#!/usr/bin/env python3
"""
Poker Tracker Web App - Version 2
With Event/Session Management and Quick Player Selection
"""

from flask import Flask, render_template, request, jsonify
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime
import json
import logging
import shutil
from pathlib import Path
import tempfile

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'poker-tracker-secret-key-2025'

# Constants
EXCEL_FILE = 'poker_tracker.xlsx'
EVENTS_FILE = 'events.json'
EVENTS_BACKUP_FILE = 'events.json.backup'
LEGACY_EVENTS_FILE = 'event_storage.json'  # For backwards compatibility with PR #1
SETTLEMENTS_FILE = 'settlements_tracking.json'
SETTLEMENTS_BACKUP_FILE = 'settlements_tracking.json.backup'
FLOAT_PRECISION_EPSILON = 0.01  # For floating point comparisons

def file_or_backup_exists(primary_path, backup_path=None):
    """
    Check if primary file or backup file exists.
    
    Args:
        primary_path: Path to primary file
        backup_path: Path to backup file (optional)
    
    Returns:
        bool: True if either file exists
    """
    return os.path.exists(primary_path) or (backup_path and os.path.exists(backup_path))

def safe_json_load(filepath, backup_filepath=None, default_value=None):
    """
    Safely load JSON file with fallback to backup and corruption handling.
    
    Args:
        filepath: Primary JSON file path
        backup_filepath: Backup file path (optional)
        default_value: Default value to return if all loads fail
    
    Returns:
        Loaded JSON data or default_value
    """
    logger.info(f"Loading JSON from {filepath}")
    
    # Try loading primary file
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r') as f:
                data = json.load(f)
            logger.info(f"Successfully loaded {filepath}")
            return data
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error in {filepath}: {e}")
        except Exception as e:
            logger.error(f"Error reading {filepath}: {e}")
    
    # Try loading backup file
    if backup_filepath and os.path.exists(backup_filepath):
        logger.warning(f"Attempting to load backup from {backup_filepath}")
        try:
            with open(backup_filepath, 'r') as f:
                data = json.load(f)
            logger.info(f"Successfully loaded backup from {backup_filepath}")
            # Restore the primary file from backup
            try:
                shutil.copy2(backup_filepath, filepath)
                logger.info(f"Restored {filepath} from backup")
            except Exception as e:
                logger.error(f"Failed to restore primary file from backup: {e}")
            return data
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error in backup {backup_filepath}: {e}")
        except Exception as e:
            logger.error(f"Error reading backup {backup_filepath}: {e}")
    
    # Return default value
    logger.warning(f"Returning default value for {filepath}")
    return default_value if default_value is not None else {}

def safe_json_save(filepath, data, backup_filepath=None):
    """
    Safely save JSON file with atomic write and backup creation.
    
    Args:
        filepath: Target JSON file path
        data: Data to save
        backup_filepath: Backup file path (optional)
    
    Returns:
        bool: True if save was successful, False otherwise
    """
    logger.info(f"Saving JSON to {filepath}")
    
    try:
        # Create backup of existing file before writing
        if backup_filepath and os.path.exists(filepath):
            try:
                shutil.copy2(filepath, backup_filepath)
                logger.info(f"Created backup at {backup_filepath}")
            except Exception as e:
                logger.warning(f"Failed to create backup: {e}")
        
        # Write to temporary file first (atomic write)
        # Use restrictive permissions for security
        dir_path = os.path.dirname(filepath) or '.'
        temp_fd, temp_path = tempfile.mkstemp(suffix='.json', dir=dir_path, text=True)
        
        try:
            # Set restrictive permissions for security
            os.chmod(temp_path, 0o600)
            
            with os.fdopen(temp_fd, 'w') as f:
                json.dump(data, f, indent=2)
            
            # Move temporary file to target location
            shutil.move(temp_path, filepath)
            logger.info(f"Successfully saved {filepath}")
            return True
        except Exception as e:
            # Clean up temp file if it exists
            try:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
            except:
                pass
            raise e
    except Exception as e:
        logger.error(f"Error saving {filepath}: {e}")
        return False

def load_settlement_payments():
    """Load settlement payment tracking from JSON with error handling"""
    return safe_json_load(
        SETTLEMENTS_FILE, 
        SETTLEMENTS_BACKUP_FILE,
        default_value={}
    )

def save_settlement_payments(payments):
    """Save settlement payment tracking to JSON with backup"""
    return safe_json_save(
        SETTLEMENTS_FILE,
        payments,
        SETTLEMENTS_BACKUP_FILE
    )

def load_events():
    """Load events list from JSON file with error handling and backwards compatibility"""
    # Check for legacy event_storage.json file from PR #1 and migrate if needed
    if not os.path.exists(EVENTS_FILE) and os.path.exists(LEGACY_EVENTS_FILE):
        logger.info(f"Migrating from {LEGACY_EVENTS_FILE} to {EVENTS_FILE}")
        try:
            with open(LEGACY_EVENTS_FILE, 'r') as f:
                events = json.load(f)
            # Save to new location using safe save
            if save_events(events):
                logger.info(f"Successfully migrated {len(events)} events")
            return events
        except Exception as e:
            logger.error(f"Failed to migrate from {LEGACY_EVENTS_FILE}: {e}")
    
    return safe_json_load(
        EVENTS_FILE,
        EVENTS_BACKUP_FILE,
        default_value=[]
    )

def save_events(events):
    """Save events list to JSON file with backup"""
    return safe_json_save(
        EVENTS_FILE,
        events,
        EVENTS_BACKUP_FILE
    )

def get_or_create_sheet(wb, sheet_name):
    """Get existing sheet or create new one"""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    
    # Create new sheet
    ws = wb.create_sheet(sheet_name)
    
    # Headers
    headers = ['Player Name', 'Phone', 'Start', 'Buy-ins', 'Day 1 End', 'Day 2 End', 'Day 3 End', 
               'Day 4 End', 'Day 5 End', 'Day 6 End', 'Day 7 End', 'P/L ($)', 'Days Played']
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color='1DA1F2', end_color='1DA1F2', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    return ws

def create_or_load_workbook():
    """Create new workbook or load existing one"""
    if os.path.exists(EXCEL_FILE):
        return openpyxl.load_workbook(EXCEL_FILE)
    
    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Events"
    ws.append(['Event List'])
    wb.save(EXCEL_FILE)
    return wb

def calculate_pl(player_data):
    """Calculate P/L based on last filled day and buy-ins"""
    start = player_data.get('start', 20)
    buyins = player_data.get('buyins', 0)
    days_played = 0
    last_value = start
    
    for day in range(1, 8):
        day_key = f'day{day}'
        if day_key in player_data and player_data[day_key]:
            try:
                days_played += 1
                last_value = float(player_data[day_key])
            except:
                pass
    
    if days_played == 0:
        return 0
    
    total_investment = start + (buyins * 20)
    pl = last_value - total_investment
    return round(pl, 2)

def calculate_settlements(players):
    """Calculate optimal settlements using greedy algorithm"""
    # Separate winners and losers
    winners = [(p['name'], p['pl']) for p in players if p['pl'] > 0]
    losers = [(p['name'], -p['pl']) for p in players if p['pl'] < 0]
    
    # Sort by amount (descending)
    winners.sort(key=lambda x: x[1], reverse=True)
    losers.sort(key=lambda x: x[1], reverse=True)
    
    settlements = []
    i, j = 0, 0
    
    while i < len(winners) and j < len(losers):
        winner_name, win_amount = winners[i]
        loser_name, loss_amount = losers[j]
        
        payment = min(win_amount, loss_amount)
        settlements.append({
            'from': loser_name,
            'to': winner_name,
            'amount': round(payment, 2)
        })
        
        winners[i] = (winner_name, win_amount - payment)
        losers[j] = (loser_name, loss_amount - payment)
        
        if winners[i][1] < FLOAT_PRECISION_EPSILON:  # Account for floating point precision
            i += 1
        if losers[j][1] < FLOAT_PRECISION_EPSILON:
            j += 1
    
    return settlements

@app.route('/')
def index():
    """Main page"""
    return render_template('index_v2.html')

@app.route('/health')
def health():
    """Health check endpoint for monitoring"""
    try:
        # Check if critical files are accessible
        checks = {
            'events_file': file_or_backup_exists(EVENTS_FILE, EVENTS_BACKUP_FILE),
            'excel_file': os.path.exists(EXCEL_FILE),
            'settlements_file': file_or_backup_exists(SETTLEMENTS_FILE, SETTLEMENTS_BACKUP_FILE)
        }
        
        # Try to load events to verify file integrity
        try:
            events = load_events()
            checks['events_loadable'] = True
            checks['event_count'] = len(events) if isinstance(events, list) else 0
        except Exception as e:
            logger.error(f"Health check: Failed to load events - {e}")
            checks['events_loadable'] = False
            checks['event_count'] = 0
        
        status = 'healthy' if all([checks['events_file'], checks['events_loadable']]) else 'degraded'
        
        return jsonify({
            'status': status,
            'timestamp': datetime.now().isoformat(),
            'checks': checks
        }), 200 if status == 'healthy' else 503
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/test')
def test():
    """Test page"""
    return render_template('test.html')

@app.route('/api/events', methods=['GET'])
def get_events():
    """Get list of all events"""
    try:
        logger.info("GET /api/events called")
        events = load_events()
        
        # Validate events data
        if not isinstance(events, list):
            logger.error(f"Invalid events data type: {type(events)}")
            events = []
        
        logger.info(f"Returning {len(events)} events")
        return jsonify({'events': events})
    except Exception as e:
        logger.error(f"Error in get_events: {e}", exc_info=True)
        return jsonify({
            'success': False, 
            'error': 'Failed to load events',
            'events': []
        }), 500

@app.route('/api/events', methods=['POST'])
def create_event():
    """Create new event"""
    try:
        logger.info("POST /api/events called")
        data = request.json
        
        if not data:
            logger.warning("No JSON data provided")
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        event_name = data.get('event_name', '').strip()
        logger.info(f"Creating event: '{event_name}'")
        
        if not event_name:
            logger.warning("Empty event name")
            return jsonify({'success': False, 'error': 'Event name required'}), 400
        
        # Load existing events
        events = load_events()
        
        # Validate events data
        if not isinstance(events, list):
            logger.error(f"Invalid events data type: {type(events)}, resetting to empty list")
            events = []
        
        # Check if event already exists
        if event_name in events:
            logger.warning(f"Event '{event_name}' already exists")
            return jsonify({'success': False, 'error': 'Event already exists'}), 400
        
        # Add new event
        events.append(event_name)
        
        # Save events with error handling
        if not save_events(events):
            logger.error("Failed to save events file")
            return jsonify({'success': False, 'error': 'Failed to save event'}), 500
        
        logger.info(f"Event '{event_name}' saved to JSON")
        
        # Create sheet in Excel
        try:
            wb = create_or_load_workbook()
            get_or_create_sheet(wb, event_name)
            wb.save(EXCEL_FILE)
            wb.close()
            logger.info(f"Excel sheet created for '{event_name}'")
        except Exception as e:
            logger.error(f"Error creating Excel sheet: {e}", exc_info=True)
            # Remove event from list since Excel creation failed
            events.remove(event_name)
            save_events(events)
            return jsonify({
                'success': False, 
                'error': 'Failed to create event workspace'
            }), 500
        
        logger.info(f"Event '{event_name}' created successfully")
        return jsonify({'success': True, 'event_name': event_name})
    except Exception as e:
        logger.error(f"Error in create_event: {e}", exc_info=True)
        return jsonify({
            'success': False, 
            'error': 'Failed to create event'
        }), 500

@app.route('/api/events/<event_name>', methods=['DELETE'])
def delete_event(event_name):
    """Delete an event"""
    try:
        logger.info(f"DELETE /api/events/{event_name} called")
        events = load_events()
        
        # Validate events data
        if not isinstance(events, list):
            logger.error(f"Invalid events data type: {type(events)}")
            return jsonify({'success': False, 'error': 'Invalid events data'}), 500
        
        if event_name not in events:
            logger.warning(f"Event '{event_name}' not found")
            return jsonify({'success': False, 'error': 'Event not found'}), 404
        
        # Remove from events list
        events.remove(event_name)
        if not save_events(events):
            logger.error("Failed to save events after deletion")
            return jsonify({'success': False, 'error': 'Failed to update events'}), 500
        
        logger.info(f"Event '{event_name}' removed from JSON")
        
        # Remove sheet from Excel
        try:
            wb = create_or_load_workbook()
            if event_name in wb.sheetnames:
                del wb[event_name]
                wb.save(EXCEL_FILE)
                logger.info(f"Excel sheet '{event_name}' deleted")
            wb.close()
        except Exception as e:
            logger.error(f"Error deleting Excel sheet: {e}", exc_info=True)
        
        # Remove settlement tracking for this event
        try:
            settlements = load_settlement_payments()
            if event_name in settlements:
                del settlements[event_name]
                save_settlement_payments(settlements)
                logger.info(f"Settlement tracking for '{event_name}' deleted")
        except Exception as e:
            logger.error(f"Error deleting settlement tracking: {e}", exc_info=True)
        
        logger.info(f"Event '{event_name}' deleted successfully")
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error in delete_event: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Failed to delete event'}), 500

@app.route('/api/data/<event_name>', methods=['GET'])
def get_event_data(event_name):
    """Get data for specific event"""
    try:
        logger.info(f"GET /api/data/{event_name} called")
        wb = create_or_load_workbook()
        
        if event_name not in wb.sheetnames:
            logger.warning(f"Event '{event_name}' not found in workbook")
            wb.close()
            return jsonify({'players': []})
        
        ws = wb[event_name]
        players = []
        
        # Read data from sheet (skip header row)
        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]
            player_data = {
                'row': row_idx,
                'name': row[0].value or '',
                'phone': row[1].value or '',
                'start': row[2].value or 20,
                'buyins': row[3].value or 0,
                'day1': row[4].value or '',
                'day2': row[5].value or '',
                'day3': row[6].value or '',
                'day4': row[7].value or '',
                'day5': row[8].value or '',
                'day6': row[9].value or '',
                'day7': row[10].value or '',
                'pl': row[11].value or 0,
                'days_played': row[12].value or 0
            }
            
            # Only include rows with names
            if player_data['name']:
                players.append(player_data)
        
        wb.close()
        logger.info(f"Loaded {len(players)} players for event '{event_name}'")
        return jsonify({'players': players})
    except Exception as e:
        logger.error(f"Error in get_event_data: {e}", exc_info=True)
        return jsonify({'players': [], 'error': 'Failed to load event data'}), 500

@app.route('/api/save/<event_name>', methods=['POST'])
def save_event_data(event_name):
    """Save data for specific event"""
    try:
        logger.info(f"POST /api/save/{event_name} called")
        data = request.json
        
        if not data:
            logger.warning("No data provided")
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        players = data.get('players', [])
        logger.info(f"Saving {len(players)} players for event '{event_name}'")
        
        wb = create_or_load_workbook()
        ws = get_or_create_sheet(wb, event_name)
        
        # Clear existing data (keep headers)
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, 14):
                ws.cell(row=row_idx, column=col_idx, value='')
        
        # Write new data
        for idx, player in enumerate(players):
            row_idx = idx + 2
            
            ws.cell(row=row_idx, column=1, value=player.get('name', ''))
            ws.cell(row=row_idx, column=2, value=player.get('phone', ''))
            ws.cell(row=row_idx, column=3, value=player.get('start', 20))
            ws.cell(row=row_idx, column=4, value=player.get('buyins', 0))  # Save buy-ins!
            
            # Day values
            for day in range(1, 8):
                day_val = player.get(f'day{day}', '')
                if day_val:
                    try:
                        ws.cell(row=row_idx, column=4+day, value=float(day_val))
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Invalid day{day} value for player {player.get('name')}: {day_val}")
                        # Leave cell empty if value is invalid
            
            # Calculate P/L and days played
            pl = calculate_pl(player)
            days_played = sum(1 for day in range(1, 8) if player.get(f'day{day}'))
            
            ws.cell(row=row_idx, column=12, value=pl)
            ws.cell(row=row_idx, column=13, value=days_played)
        
        wb.save(EXCEL_FILE)
        wb.close()
        logger.info(f"Successfully saved data for event '{event_name}'")
        
        return jsonify({'success': True, 'message': 'Data saved successfully'})
    except Exception as e:
        logger.error(f"Error in save_event_data: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Failed to save data'}), 500

@app.route('/api/settlements/<event_name>', methods=['GET'])
def get_event_settlements(event_name):
    """Calculate settlements for specific event"""
    try:
        logger.info(f"GET /api/settlements/{event_name} called")
        wb = create_or_load_workbook()
        
        if event_name not in wb.sheetnames:
            logger.warning(f"Event '{event_name}' not found in workbook")
            wb.close()
            return jsonify({'settlements': []})
        
        ws = wb[event_name]
        players = []
        
        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]
            name = row[0].value
            pl = row[11].value or 0
            
            if name and pl != 0:
                players.append({'name': name, 'pl': float(pl)})
        
        wb.close()
        
        if not players:
            logger.info(f"No player data found for event '{event_name}'")
            return jsonify({'settlements': [], 'message': 'No player data found'})
        
        settlements = calculate_settlements(players)
        
        # Load payment status for this event
        payment_tracking = load_settlement_payments()
        event_payments = payment_tracking.get(event_name, {})
        
        # Add payment status to each settlement
        for settlement in settlements:
            settlement_key = f"{settlement['from']}→{settlement['to']}"
            settlement['paid'] = event_payments.get(settlement_key, False)
        
        logger.info(f"Calculated {len(settlements)} settlements for event '{event_name}'")
        return jsonify({
            'settlements': settlements,
            'total_winners': sum(p['pl'] for p in players if p['pl'] > 0),
            'total_losers': abs(sum(p['pl'] for p in players if p['pl'] < 0))
        })
    except Exception as e:
        logger.error(f"Error in get_event_settlements: {e}", exc_info=True)
        return jsonify({
            'settlements': [], 
            'error': 'Failed to calculate settlements'
        }), 500

@app.route('/api/settlements/<event_name>/mark_paid', methods=['POST'])
def mark_settlement_paid(event_name):
    """Mark a settlement as paid or unpaid"""
    try:
        logger.info(f"POST /api/settlements/{event_name}/mark_paid called")
        data = request.json
        
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        from_player = data.get('from')
        to_player = data.get('to')
        paid = data.get('paid', True)
        
        if not from_player or not to_player:
            return jsonify({'success': False, 'error': 'Missing player information'}), 400
        
        settlement_key = f"{from_player}→{to_player}"
        
        # Load existing payment tracking
        payment_tracking = load_settlement_payments()
        
        # Initialize event tracking if not exists
        if event_name not in payment_tracking:
            payment_tracking[event_name] = {}
        
        # Update payment status
        payment_tracking[event_name][settlement_key] = paid
        
        # Save back to file
        if not save_settlement_payments(payment_tracking):
            logger.error("Failed to save settlement payment status")
            return jsonify({'success': False, 'error': 'Failed to save payment status'}), 500
        
        logger.info(f"Settlement {settlement_key} marked as {'paid' if paid else 'unpaid'}")
        return jsonify({
            'success': True,
            'message': f"Settlement marked as {'paid' if paid else 'unpaid'}"
        })
    except Exception as e:
        logger.error(f"Error in mark_settlement_paid: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Failed to update payment status'}), 500

@app.route('/api/clear/<event_name>', methods=['POST'])
def clear_event_data(event_name):
    """Clear data for specific event"""
    try:
        logger.info(f"POST /api/clear/{event_name} called")
        wb = create_or_load_workbook()
        
        if event_name in wb.sheetnames:
            ws = wb[event_name]
            # Keep headers, clear data rows
            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, 14):
                    ws.cell(row=row_idx, column=col_idx, value='')
            wb.save(EXCEL_FILE)
            logger.info(f"Cleared data for event '{event_name}'")
        
        wb.close()
        return jsonify({'success': True, 'message': f'Event "{event_name}" cleared successfully'})
    except Exception as e:
        logger.error(f"Error in clear_event_data: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Failed to clear event data'}), 500

def initialize_app():
    """Initialize application state on startup or wake from sleep"""
    logger.info("=" * 60)
    logger.info("Initializing Poker Tracker Application")
    logger.info("=" * 60)
    
    try:
        # Ensure Excel file exists
        wb = create_or_load_workbook()
        wb.close()
        logger.info("✓ Excel workbook initialized")
    except Exception as e:
        logger.error(f"✗ Failed to initialize Excel workbook: {e}")
    
    try:
        # Validate events.json
        events = load_events()
        if isinstance(events, list):
            logger.info(f"✓ Events file loaded successfully ({len(events)} events)")
        else:
            logger.warning(f"✗ Events file has invalid format, resetting")
            save_events([])
    except Exception as e:
        logger.error(f"✗ Failed to load events: {e}")
        try:
            save_events([])
            logger.info("✓ Created new events file")
        except Exception as e2:
            logger.error(f"✗ Failed to create events file: {e2}")
    
    try:
        # Validate settlements.json
        settlements = load_settlement_payments()
        if isinstance(settlements, dict):
            logger.info(f"✓ Settlements file loaded successfully ({len(settlements)} events)")
        else:
            logger.warning(f"✗ Settlements file has invalid format, resetting")
            save_settlement_payments({})
    except Exception as e:
        logger.error(f"✗ Failed to load settlements: {e}")
        try:
            save_settlement_payments({})
            logger.info("✓ Created new settlements file")
        except Exception as e2:
            logger.error(f"✗ Failed to create settlements file: {e2}")
    
    logger.info("=" * 60)
    logger.info("Application initialization complete")
    logger.info("=" * 60)

if __name__ == '__main__':
    # Initialize application state
    initialize_app()
    
    # Get port from environment variable (for deployment) or use 5001 for local
    port = int(os.environ.get('PORT', 5001))
    
    # Check if running in production
    is_production = os.environ.get('RENDER', False)
    
    logger.info("")
    logger.info("🎰 Poker Tracker Web App v2 - Event Management")
    logger.info("=" * 50)
    if is_production:
        logger.info("🌐 Production Mode - Deployed on Render")
    else:
        logger.info("🌐 Development Mode - Running Locally")
        logger.info(f"📱 Open on your phone: http://YOUR_COMPUTER_IP:{port}")
        logger.info(f"💻 Open on this computer: http://localhost:{port}")
    logger.info("")
    logger.info("✅ Ready! Press Ctrl+C to stop")
    logger.info("=" * 50)
    logger.info("")
    
    app.run(host='0.0.0.0', port=port, debug=not is_production)
