#!/usr/bin/env python3
"""
Poker Tracker Web App
Mobile-friendly web application for tracking poker game settlements
"""

from flask import Flask, render_template, request, jsonify
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from datetime import datetime

app = Flask(__name__)
app.config['SECRET_KEY'] = 'poker-tracker-secret-key-2025'

EXCEL_FILE = 'poker_tracker.xlsx'

def create_or_load_workbook():
    """Create new workbook or load existing one"""
    if os.path.exists(EXCEL_FILE):
        return openpyxl.load_workbook(EXCEL_FILE)
    
    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Poker Tracker"
    
    # Headers
    headers = ['Player Name', 'Phone', 'Start', 'Day 1 End', 'Day 2 End', 'Day 3 End', 
               'Day 4 End', 'Day 5 End', 'Day 6 End', 'Day 7 End', 'P/L ($)', 'Days Played', 'Notes']
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color='1DA1F2', end_color='1DA1F2', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add 10 empty player rows
    for i in range(10):
        ws.append([''] * 13)
        ws.cell(row=i+2, column=3, value=20)  # Default $20 start
    
    wb.save(EXCEL_FILE)
    return wb

def calculate_pl(row_data):
    """Calculate P/L based on last filled day"""
    start = row_data.get('start', 20)
    days_played = 0
    last_value = start
    
    for day in range(1, 8):
        day_key = f'day{day}'
        if day_key in row_data and row_data[day_key]:
            days_played += 1
            last_value = float(row_data[day_key])
    
    if days_played == 0:
        return 0
    
    total_investment = start + (days_played - 1) * 20
    pl = last_value - total_investment
    return round(pl, 2)

def calculate_settlements(players):
    """Calculate optimal settlements using greedy algorithm"""
    # Separate winners and losers
    winners = [(p['name'], p['pl']) for p in players if p['pl'] > 0]
    losers = [(p['name'], -p['pl']) for p in players if p['pl'] < 0]
    
    # Sort by amount (descending)
    winners.sort(key=lambda x: x[1], reverse=True)
    losers.sort(key=lambda x: x[1], reverse=True)
    
    settlements = []
    i, j = 0, 0
    
    while i < len(winners) and j < len(losers):
        winner_name, win_amount = winners[i]
        loser_name, loss_amount = losers[j]
        
        payment = min(win_amount, loss_amount)
        settlements.append({
            'from': loser_name,
            'to': winner_name,
            'amount': round(payment, 2)
        })
        
        winners[i] = (winner_name, win_amount - payment)
        losers[j] = (loser_name, loss_amount - payment)
        
        if winners[i][1] == 0:
            i += 1
        if losers[j][1] == 0:
            j += 1
    
    return settlements

@app.route('/')
def index():
    """Main page"""
    return render_template('index.html')

@app.route('/api/data', methods=['GET'])
def get_data():
    """Get all poker tracker data"""
    wb = create_or_load_workbook()
    ws = wb.active
    
    players = []
    for row_idx in range(2, 12):  # Rows 2-11 (10 players)
        row = ws[row_idx]
        player_data = {
            'row': row_idx,
            'name': row[0].value or '',
            'phone': row[1].value or '',
            'start': row[2].value or 20,
            'day1': row[3].value or '',
            'day2': row[4].value or '',
            'day3': row[5].value or '',
            'day4': row[6].value or '',
            'day5': row[7].value or '',
            'day6': row[8].value or '',
            'day7': row[9].value or '',
            'pl': row[10].value or 0,
            'days_played': row[11].value or 0,
            'notes': row[12].value or ''
        }
        if player_data['name']:  # Only include rows with names
            players.append(player_data)
    
    wb.close()
    return jsonify({'players': players})

@app.route('/api/save', methods=['POST'])
def save_data():
    """Save poker tracker data"""
    data = request.json
    players = data.get('players', [])
    
    wb = create_or_load_workbook()
    ws = wb.active
    
    # Clear existing data (keep headers)
    for row_idx in range(2, 12):
        for col_idx in range(1, 14):
            ws.cell(row=row_idx, column=col_idx, value='')
    
    # Write new data
    for idx, player in enumerate(players):
        row_idx = idx + 2
        
        ws.cell(row=row_idx, column=1, value=player.get('name', ''))
        ws.cell(row=row_idx, column=2, value=player.get('phone', ''))
        ws.cell(row=row_idx, column=3, value=player.get('start', 20))
        
        # Day values
        for day in range(1, 8):
            day_val = player.get(f'day{day}', '')
            if day_val:
                ws.cell(row=row_idx, column=3+day, value=float(day_val))
        
        # Calculate P/L and days played
        pl = calculate_pl(player)
        days_played = sum(1 for day in range(1, 8) if player.get(f'day{day}'))
        
        ws.cell(row=row_idx, column=11, value=pl)
        ws.cell(row=row_idx, column=12, value=days_played)
        ws.cell(row=row_idx, column=13, value=player.get('notes', ''))
    
    wb.save(EXCEL_FILE)
    wb.close()
    
    return jsonify({'success': True, 'message': 'Data saved successfully'})

@app.route('/api/settlements', methods=['GET'])
def get_settlements():
    """Calculate and return settlement instructions"""
    wb = create_or_load_workbook()
    ws = wb.active
    
    players = []
    for row_idx in range(2, 12):
        row = ws[row_idx]
        name = row[0].value
        pl = row[10].value or 0
        
        if name and pl != 0:
            players.append({'name': name, 'pl': float(pl)})
    
    wb.close()
    
    if not players:
        return jsonify({'settlements': [], 'message': 'No player data found'})
    
    settlements = calculate_settlements(players)
    
    return jsonify({
        'settlements': settlements,
        'total_winners': sum(p['pl'] for p in players if p['pl'] > 0),
        'total_losers': abs(sum(p['pl'] for p in players if p['pl'] < 0))
    })

@app.route('/api/clear', methods=['POST'])
def clear_data():
    """Clear all data (reset)"""
    if os.path.exists(EXCEL_FILE):
        os.remove(EXCEL_FILE)
    create_or_load_workbook()
    return jsonify({'success': True, 'message': 'Data cleared successfully'})

if __name__ == '__main__':
    import os
    
    create_or_load_workbook()
    
    # Get port from environment variable (for deployment) or use 5001 for local
    port = int(os.environ.get('PORT', 5001))
    
    # Check if running in production
    is_production = os.environ.get('RENDER', False)
    
    print("\n🎰 Poker Tracker Web App")
    print("=" * 50)
    if is_production:
        print("🌐 Production Mode - Deployed on Render")
    else:
        print("🌐 Development Mode - Running Locally")
        print("📱 Open on your phone:")
        print(f"   http://YOUR_COMPUTER_IP:{port}")
        print("\n💻 Open on this computer:")
        print(f"   http://localhost:{port}")
    print("\n✅ Ready! Press Ctrl+C to stop")
    print("=" * 50 + "\n")
    
    app.run(host='0.0.0.0', port=port, debug=not is_production)
